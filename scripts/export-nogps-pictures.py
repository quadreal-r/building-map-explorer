#!/usr/bin/env python3
"""Export filenames from the NoGPS Pictures folder to Excel."""
from __future__ import annotations

import importlib.util
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_SCRIPT = Path(__file__).with_name("export-rtu-picture-duplicates.py")
_spec = importlib.util.spec_from_file_location("export_rtu_picture_duplicates", _SCRIPT)
_mod = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_mod)

PICTURES_DIR = _mod.PICTURES_DIR
NOGPS_DIR = PICTURES_DIR / "NoGPS Pictures"
OUTPUT_PATH = NOGPS_DIR / "NoGPS-Pictures-List.xlsx"
parse_rtu_key = _mod.parse_rtu_key
read_image_metadata = _mod.read_image_metadata
kb = _mod.kb
autosize_columns = _mod.autosize_columns
style_header = _mod.style_header


def collect_nogps_images() -> list[dict]:
    rows: list[dict] = []
    for path in sorted(NOGPS_DIR.glob("*.jpg")):
        parsed = parse_rtu_key(path.name)
        meta = read_image_metadata(path)
        rows.append(
            {
                "path": path,
                "name": path.name,
                "size": path.stat().st_size,
                "building": parsed[0] if parsed else "",
                "rtu_key": parsed[1] if parsed else "",
                "unit": parsed[2] if parsed else "",
                "parsed": parsed is not None,
                **meta,
            }
        )
    return rows


def main() -> None:
    if not NOGPS_DIR.is_dir():
        raise SystemExit(f"Folder not found: {NOGPS_DIR}")

    images = collect_nogps_images()
    wb = Workbook()
    ws = wb.active
    ws.title = "No GPS Pictures"
    ws.append(
        [
            "#",
            "Filename",
            "Building",
            "RTU key",
            "Unit",
            "Size (KB)",
            "Has EXIF",
            "Capture time",
            "GPS note",
        ]
    )
    for i, img in enumerate(images, 1):
        ws.append(
            [
                i,
                img["name"],
                img["building"],
                img["rtu_key"],
                img["unit"],
                kb(img["size"]),
                "Yes" if img["has_exif"] else "No",
                img["datetime"],
                img["gps_note"] or "No GPS coordinates",
            ]
        )
    style_header(ws)
    autosize_columns(ws)

    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["NoGPS Pictures export"])
    ws_sum.append(["Folder", str(NOGPS_DIR)])
    ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_sum.append([])
    ws_sum.append(["Total files", len(images)])
    ws_sum.append(["Parsed filenames", sum(1 for i in images if i["parsed"])])
    ws_sum.append(["Unparsed filenames", sum(1 for i in images if not i["parsed"])])
    buildings = sorted({i["building"] for i in images if i["building"]})
    ws_sum.append(["Buildings", len(buildings)])
    ws_sum.append([])
    ws_sum.append(["Building", "File count"])
    for building in buildings:
        count = sum(1 for i in images if i["building"] == building)
        ws_sum.append([building, count])
    autosize_columns(ws_sum)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"  Files: {len(images)}")


if __name__ == "__main__":
    main()
