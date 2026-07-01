#!/usr/bin/env python3
"""Export RTU picture duplicate analysis to Excel."""
from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from PIL import Image
from PIL.ExifTags import IFD
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PICTURES_DIR = Path(
    r"C:\Users\Robert\OneDrive - Quadreal Property Group"
    r"\#OI-Industrial East - @(RTU) Roof Top Units (All Industrial)"
    r"\RTUs per Building\_RTU-Pictures-All"
)
OUTPUT_PATH = PICTURES_DIR / "RTU-Picture-Duplicates-Report.xlsx"

RTU_KEY_RE = re.compile(r"^(\d+)-RTU-([\d]+[A-Za-z]?)")


def parse_rtu_key(name: str) -> tuple[str, str, str] | None:
    m = RTU_KEY_RE.match(name)
    if not m:
        return None
    building, unit = m.group(1), m.group(2)
    return building, f"{building}-RTU-{unit}", unit


def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def kb(size: int) -> float:
    return round(size / 1024, 1)


def _rational_to_float(value) -> float:
    if isinstance(value, tuple):
        return float(value[0]) / float(value[1])
    return float(value)


def _dms_to_decimal(dms) -> float:
    degrees, minutes, seconds = dms
    return (
        _rational_to_float(degrees)
        + _rational_to_float(minutes) / 60
        + _rational_to_float(seconds) / 3600
    )


def _exif_datetime(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8", errors="replace")
    text = str(raw).strip()
    for fmt in ("%Y:%m:%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text


def read_image_metadata(path: Path) -> dict:
    """Read GPS and capture time from EXIF when present."""
    result = {
        "has_exif": False,
        "has_gps": False,
        "lat": None,
        "lng": None,
        "datetime": "",
        "gps_note": "No EXIF",
    }
    try:
        with Image.open(path) as img:
            exif = img.getexif()
            if not exif:
                return result
            result["has_exif"] = True
            result["gps_note"] = "EXIF present, no GPS"

            for tag in (0x9003, 0x9004, 0x0132):  # DateTimeOriginal, DateTimeDigitized, DateTime
                if tag in exif:
                    result["datetime"] = _exif_datetime(exif.get(tag))
                    break

            gps = exif.get_ifd(IFD.GPSInfo)
            if not gps:
                return result

            lat, lat_ref = gps.get(2), gps.get(1)
            lon, lon_ref = gps.get(4), gps.get(3)
            if lat is None or lon is None:
                return result

            lat_dec = _dms_to_decimal(lat)
            lon_dec = _dms_to_decimal(lon)
            lat_ref_s = lat_ref.decode() if isinstance(lat_ref, bytes) else str(lat_ref or "")
            lon_ref_s = lon_ref.decode() if isinstance(lon_ref, bytes) else str(lon_ref or "")
            if lat_ref_s.upper() == "S":
                lat_dec = -lat_dec
            if lon_ref_s.upper() == "W":
                lon_dec = -lon_dec

            if not (-90 <= lat_dec <= 90 and -180 <= lon_dec <= 180):
                result["gps_note"] = "Invalid GPS coordinates"
                return result

            result.update(
                has_gps=True,
                lat=round(lat_dec, 7),
                lng=round(lon_dec, 7),
                gps_note="",
            )
    except Exception as exc:
        result["gps_note"] = f"Read error: {exc}"
    return result


def autosize_columns(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = min(width + 2, 60)


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="366092")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        if cell.value:
            cell.fill = fill
            cell.font = font


def collect_images() -> list[dict]:
    rows: list[dict] = []
    paths = sorted(PICTURES_DIR.rglob("*.jpg"))
    for i, path in enumerate(paths, 1):
        if i % 200 == 0:
            print(f"  Reading EXIF {i}/{len(paths)}...")
        parsed = parse_rtu_key(path.name)
        meta = read_image_metadata(path)
        rows.append(
            {
                "path": path,
                "name": path.name,
                "size": path.stat().st_size,
                "building": parsed[0] if parsed else "",
                "rtu_key": parsed[1] if parsed else "",
                "unit": parsed[2] if parsed else "",
                "parsed": parsed is not None,
                **meta,
            }
        )
    return rows


def main() -> None:
    images = collect_images()
    print(f"Scanning {len(images)} images...")

    # Hash cache
    hashes: dict[Path, str] = {}

    def get_hash(path: Path) -> str:
        if path not in hashes:
            hashes[path] = file_hash(path)
        return hashes[path]

    # Within-RTU: same rtu_key + size
    within_groups: dict[tuple[str, int], list[dict]] = defaultdict(list)
    for img in images:
        if not img["parsed"]:
            continue
        within_groups[(img["rtu_key"], img["size"])].append(img)

    within_dupes = {k: v for k, v in within_groups.items() if len(v) > 1}

    # Cross-RTU: same building + size, 2+ distinct rtu keys
    cross_groups: dict[tuple[str, int], list[dict]] = defaultdict(list)
    for img in images:
        if not img["parsed"]:
            continue
        cross_groups[(img["building"], img["size"])].append(img)

    cross_dupes: list[tuple[tuple[str, int], list[dict]]] = []
    for key, group in cross_groups.items():
        rtu_keys = {g["rtu_key"] for g in group}
        if len(rtu_keys) >= 2:
            cross_dupes.append((key, group))
    cross_dupes.sort(key=lambda x: (x[0][0], x[0][1]))

    wb = Workbook()

    # --- Summary ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["RTU Picture Duplicate Report"])
    ws_sum.append(["Source folder", str(PICTURES_DIR)])
    ws_sum.append([])
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total JPG files", len(images)])
    ws_sum.append(["Unparsed filenames", sum(1 for i in images if not i["parsed"])])
    ws_sum.append(["Within-RTU duplicate groups (same RTU + size)", len(within_dupes)])
    ws_sum.append(["Within-RTU extra files", sum(len(v) - 1 for v in within_dupes.values())])
    ws_sum.append(["Cross-RTU duplicate groups (same building + size)", len(cross_dupes)])
    cross_files = sum(len(g) for _, g in cross_dupes)
    ws_sum.append(["Cross-RTU files involved", cross_files])
    no_gps = [i for i in images if not i["has_gps"]]
    with_gps = [i for i in images if i["has_gps"]]
    ws_sum.append(["Pictures with GPS", len(with_gps)])
    ws_sum.append(["Pictures without GPS", len(no_gps)])
    ws_sum.append([])
    ws_sum.append(["Notes"])
    ws_sum.append(["Within-RTU", "Same building-RTU unit and file size (e.g. 6150-RTU-17 (1) vs (7)(2024))"])
    ws_sum.append(["Cross-RTU", "Same building and file size but different RTU units (building 6150 only)"])
    ws_sum.append(["No GPS", "JPEG files with no usable GPS coordinates in EXIF"])
    ws_sum["A1"].font = Font(bold=True, size=14)

    # --- Cross-RTU ---
    ws_cross = wb.create_sheet("Cross-RTU")
    ws_cross.append(
        [
            "Group",
            "Building",
            "Size (bytes)",
            "Size (KB)",
            "RTU units involved",
            "File count",
            "Filename",
            "RTU key",
            "Content identical in group",
        ]
    )
    for gi, ((building, size), group) in enumerate(cross_dupes, 1):
        group_hashes = {get_hash(g["path"]) for g in group}
        all_identical = len(group_hashes) == 1
        rtus = ", ".join(sorted({g["rtu_key"] for g in group}))
        sorted_group = sorted(group, key=lambda g: (g["rtu_key"], g["name"]))
        for i, g in enumerate(sorted_group):
            ws_cross.append(
                [
                    gi if i == 0 else "",
                    building if i == 0 else "",
                    size if i == 0 else "",
                    kb(size) if i == 0 else "",
                    rtus if i == 0 else "",
                    len(group) if i == 0 else "",
                    g["name"],
                    g["rtu_key"],
                    "Yes" if all_identical else "No - review",
                ]
            )
    style_header(ws_cross)
    autosize_columns(ws_cross)

    # --- Cross-RTU mapping (RTU-00 anchors) ---
    ws_map = wb.create_sheet("Cross-RTU RTU-00 map")
    ws_map.append(
        [
            "Size (KB)",
            "RTU-00 anchor file(s)",
            "Copy filed under other RTU",
            "Other RTU key",
            "SHA256 (first 16)",
        ]
    )
    hash_to_images: dict[str, list[dict]] = defaultdict(list)
    b6150 = [i for i in images if i["building"] == "6150"]
    for img in b6150:
        hash_to_images[get_hash(img["path"])].append(img)

    map_rows = 0
    for digest, group in sorted(
        hash_to_images.items(),
        key=lambda x: (-len(x[1]), x[1][0]["name"]),
    ):
        if len(group) < 2:
            continue
        rtu_keys = {g["rtu_key"] for g in group}
        if len(rtu_keys) < 2:
            continue
        rtu00 = sorted(g["name"] for g in group if g["rtu_key"] == "6150-RTU-00")
        others = sorted(
            (g["name"], g["rtu_key"])
            for g in group
            if g["rtu_key"] != "6150-RTU-00"
        )
        if not rtu00 or not others:
            continue
        anchor = ", ".join(rtu00)
        size_kb = kb(group[0]["size"])
        for j, (name, rtu_key) in enumerate(others):
            ws_map.append(
                [
                    size_kb if j == 0 else "",
                    anchor if j == 0 else "",
                    name,
                    rtu_key,
                    digest[:16] if j == 0 else "",
                ]
            )
            map_rows += 1
    style_header(ws_map)
    autosize_columns(ws_map)

    # --- Within-RTU ---
    ws_within = wb.create_sheet("Within-RTU")
    ws_within.append(
        [
            "Group",
            "RTU key",
            "Size (bytes)",
            "Size (KB)",
            "File count",
            "Filename",
            "Content identical in group",
        ]
    )
    for gi, ((rtu_key, size), group) in enumerate(
        sorted(within_dupes.items(), key=lambda x: (x[0][0], x[0][1])), 1
    ):
        group_hashes = {get_hash(g["path"]) for g in group}
        all_identical = len(group_hashes) == 1
        sorted_group = sorted(group, key=lambda g: g["name"])
        for i, g in enumerate(sorted_group):
            ws_within.append(
                [
                    gi if i == 0 else "",
                    rtu_key if i == 0 else "",
                    size if i == 0 else "",
                    kb(size) if i == 0 else "",
                    len(group) if i == 0 else "",
                    g["name"],
                    "Yes" if all_identical else "No - review",
                ]
            )
    style_header(ws_within)
    autosize_columns(ws_within)

    # --- No GPS ---
    ws_nogps = wb.create_sheet("No GPS")
    ws_nogps.append(
        [
            "Filename",
            "Building",
            "RTU key",
            "Size (KB)",
            "Has EXIF",
            "Capture time",
            "Reason",
        ]
    )
    for img in sorted(no_gps, key=lambda i: (i["building"], i["rtu_key"], i["name"])):
        ws_nogps.append(
            [
                img["name"],
                img["building"],
                img["rtu_key"],
                kb(img["size"]),
                "Yes" if img["has_exif"] else "No",
                img["datetime"],
                img["gps_note"] or "No GPS coordinates",
            ]
        )
    style_header(ws_nogps)
    autosize_columns(ws_nogps)

    # --- Unparsed ---
    unparsed = [i for i in images if not i["parsed"]]
    if unparsed:
        ws_bad = wb.create_sheet("Unparsed")
        ws_bad.append(["Filename", "Size (KB)", "Has GPS"])
        for g in unparsed:
            ws_bad.append([g["name"], kb(g["size"]), "Yes" if g["has_gps"] else "No"])
        style_header(ws_bad)
        autosize_columns(ws_bad)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Within-RTU groups: {len(within_dupes)}")
    print(f"  Cross-RTU groups: {len(cross_dupes)}")
    print(f"  No GPS: {len(no_gps)} / {len(images)}")


if __name__ == "__main__":
    main()
